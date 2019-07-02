import logging
from normality import stringify
from followthemoney import model
from openpyxl import load_workbook
from xml.etree.ElementTree import ParseError

from ingestors.ingestor import Ingestor
from ingestors.support.table import TableSupport
from ingestors.support.ooxml import OOXMLSupport
from ingestors.exc import ProcessingException

log = logging.getLogger(__name__)


class ExcelXMLIngestor(Ingestor, TableSupport, OOXMLSupport):
    MIME_TYPES = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # noqa
        'application/vnd.openxmlformats-officedocument.spreadsheetml.template',  # noqa
    ]
    EXTENSIONS = [
        'xlsx',
        'xlsm',
        'xltx',
        'xltm'
    ]
    SCORE = 7

    def generate_rows(self, sheet):
        for row in sheet.rows:
            try:
                yield [stringify(c.value) for c in row]
            except (ValueError, OverflowError, ParseError) as ve:
                log.warning("Failed to read Excel row: %s", ve)

    def ingest(self, file_path, entity):
        entity.schema = model.get('Workbook')
        self.ooxml_extract_metadata(file_path, entity)
        try:
            book = load_workbook(file_path, read_only=True)
        except Exception as err:
            raise ProcessingException('Invalid Excel file: %s' % err)

        try:
            for name in book.sheetnames:
                table = self.manager.make_entity('Table', parent=entity)
                table.make_id(entity.id, name)
                table.set('title', name)
                self.emit_row_tuples(table, self.generate_rows(book[name]))
                self.manager.emit_entity(table)
        except Exception as err:
            raise ProcessingException('Cannot read Excel file: %s' % err) from err  # noqa
        finally:
            book.close()

    @classmethod
    def match(cls, file_path, entity):
        score = super(ExcelXMLIngestor, cls).match(file_path, entity)
        if score <= 0 and cls.inspect_ooxml_manifest(file_path):
            score = cls.SCORE * 2
        return score
